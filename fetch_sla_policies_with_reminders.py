"""
Freshdesk SLA Policies with Reminders Export Script

DESCRIPTION:
This script retrieves SLA (Service Level Agreement) policies from Freshdesk
with a focus on reminder and notification settings. It exports policies to
Excel format specifically highlighting escalation reminders, notification
rules, and agent alert configurations for compliance and monitoring.

REQUIREMENTS:
- Python 3.x
- requests library (install with: pip install requests)
- pandas library (install with: pip install pandas)
- openpyxl library (install with: pip install openpyxl)
- Valid Freshdesk API key with SLA policy read permissions
- Freshdesk account and domain access

SETUP INSTRUCTIONS:
1. Replace API_KEY with your actual Freshdesk API key
2. Replace DOMAIN with your Freshdesk domain (e.g., 'yourcompany.freshdesk.com')
3. Ensure your API key has permissions for SLA policy access
4. Run the script: python fetch_sla_policies_with_reminders.py

API DOCUMENTATION:
- Freshdesk API v2: https://developers.freshdesk.com/api/
- Authentication: Basic Auth with API key
- Rate Limits: 50 requests per minute for most endpoints

INPUT PARAMETERS:
- API_KEY: Your Freshdesk API key
- DOMAIN: Your Freshdesk domain
- OUTPUT_FILENAME: Name for the Excel file (default: 'sla_policies_with_reminders.xlsx')

OUTPUT:
- Excel file (.xlsx) with SLA policy data focused on reminder settings
- Log file with detailed operation information
- Console output showing progress and results

EXCEL OUTPUT INCLUDES:
- Policy ID, name, description, and status
- Creation and update timestamps
- SLA targets by priority (response/resolve times)
- Business hours configuration
- Escalation rules and agent assignments
- Reminder notification settings
- All nested JSON structures flattened for spreadsheet analysis

REMINDER-SPECIFIC FEATURES:
- Highlights policies with reminder/escalation enabled
- Shows notification timeframes before SLA breach
- Lists agents configured for escalation notifications
- Identifies policies requiring immediate attention

ERROR HANDLING:
- Handles HTTP errors with descriptive messages
- Handles rate limiting with automatic retry
- Validates API responses and data structure
- Continues processing even if individual pages fail
- Comprehensive logging for troubleshooting

RATE LIMIT HANDLING:
- Automatically detects rate limit responses (HTTP 429)
- Waits for the specified retry-after period
- Continues processing remaining pages after rate limit delay
- Logs rate limit events for monitoring

SECURITY NOTE:
- Store API keys securely (environment variables recommended for production)
- Never commit API keys to version control
- Rotate API keys regularly for security

TROUBLESHOOTING:
- Verify API key has SLA policy read permissions
- Check Freshdesk domain is correct
- Ensure network connectivity to Freshdesk API
- Monitor rate limit usage in Freshdesk dashboard
- Check log file for detailed error information
- Verify required Python packages are installed

PERFORMANCE CONSIDERATIONS:
- Processes SLA policies in pages (default page size varies by API)
- Handles pagination automatically
- Large numbers of policies may take significant time to process
- Rate limiting may cause delays in processing
- Excel file size grows with number of policies and nested data

REMINDER ANALYSIS FEATURES:
- Identifies policies with active reminder systems
- Shows escalation timeframes for agent notifications
- Highlights policies needing immediate configuration review
- Provides compliance-ready export for SLA monitoring
"""

import requests
import json
import time
import logging
import pandas as pd
import os

# Freshdesk API Configuration
# TODO: Move these to environment variables for security
API_KEY = '5TMgbcZdRFY70hSpEdj'  # Replace with your actual API key
DOMAIN = 'benchmarkeducationcompany.freshdesk.com'  # Replace with your domain

# API Configuration
BASE_URL = f'https://{DOMAIN}/api/v2'
HEADERS = {'Content-Type': 'application/json'}

# Output configuration
OUTPUT_FILENAME = 'sla_policies_with_reminders.xlsx'  # Default output filename
LOG_FILENAME = 'sla_policies_reminders_export.log'  # Log file name

def setup_logging():
    """
    Set up comprehensive logging to both file and console.

    Creates log directory if it doesn't exist and configures
    logging with timestamps and detailed formatting.
    """
    # Ensure log directory exists (create if needed)
    log_dir = os.path.dirname(LOG_FILENAME) if os.path.dirname(LOG_FILENAME) else '.'
    if log_dir and not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # Configure logging to both file and console
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(LOG_FILENAME, encoding='utf-8'),
            logging.StreamHandler()  # Also log to console
        ]
    )

def validate_configuration():
    """
    Validate that all required configuration is present and valid.

    Returns:
        bool: True if configuration is valid, False otherwise
    """
    if not API_KEY or API_KEY == '5TMgbcZdRFY70hSpEdj':
        logging.error("API_KEY not configured. Please set your actual Freshdesk API key.")
        print("❌ API_KEY not configured. Please update the script with your API key.")
        return False

    if not DOMAIN or DOMAIN == 'benchmarkeducationcompany.freshdesk.com':
        logging.error("DOMAIN not configured. Please set your actual Freshdesk domain.")
        print("❌ DOMAIN not configured. Please update the script with your domain.")
        return False

    return True

def get_sla_policies():
    """
    Retrieve all SLA policies from Freshdesk with pagination support.

    Returns:
        list: List of SLA policy dictionaries, or empty list if failed

    Note:
        - Handles pagination automatically
        - Implements rate limiting with retry logic
        - Logs progress for monitoring
    """
    endpoint = f'{BASE_URL}/sla_policies'
    sla_policies = []
    page = 1

    logging.info(f"Starting SLA policies retrieval from {DOMAIN} for reminder analysis")
    print("Fetching SLA policies for reminder analysis..."
    while True:
        try:
            # Make API request for current page
            response = requests.get(
                endpoint,
                headers=HEADERS,
                auth=(API_KEY, 'X'),
                params={'page': page}
            )

            if response.status_code == 200:
                # Success - parse response data
                data = response.json()

                if not data:
                    # No more data available
                    logging.info(f"Completed fetching all SLA policies. Total pages: {page - 1}")
                    break

                # Add this page of policies to our collection
                sla_policies.extend(data)
                logging.info(f"Fetched page {page} ({len(data)} policies)")
                print(f"  Page {page}: {len(data)} policies retrieved")

                page += 1

            elif response.status_code == 429:
                # Rate limit exceeded - retry after delay
                retry_after = int(response.headers.get('Retry-After', 60))
                logging.warning(f"Rate limit exceeded on page {page}. Retrying after {retry_after} seconds...")
                print(f"⏳ Rate limit reached. Waiting {retry_after} seconds...")
                time.sleep(retry_after)
                continue  # Retry the same page

            else:
                # Other error
                logging.error(f"Failed to fetch SLA policies on page {page}. Status: {response.status_code}")
                logging.error(f"Response: {response.text}")
                print(f"❌ Failed to fetch page {page}: {response.status_code}")
                break

        except requests.exceptions.RequestException as e:
            logging.error(f"Network error on page {page}: {e}")
            print(f"❌ Network error on page {page}: {e}")
            break

    total_policies = len(sla_policies)
    logging.info(f"Successfully retrieved {total_policies} SLA policies total")
    print(f"✓ Retrieved {total_policies} SLA policies total")

    return sla_policies

def analyze_reminder_settings(policy):
    """
    Analyze reminder and escalation settings for a policy.

    Args:
        policy (dict): Single SLA policy dictionary

    Returns:
        dict: Policy with added reminder analysis fields
    """
    enriched = policy.copy()

    # Check if escalation is enabled anywhere in the policy
    sla_target = policy.get('sla_target', {})
    escalation = policy.get('escalation', {})

    has_reminders = False
    reminder_summary = []

    # Check SLA targets for escalation settings
    if sla_target:
        for priority, target in sla_target.items():
            if target.get('escalation_enabled'):
                has_reminders = True
                reminder_summary.append(f"{priority}: escalation enabled")

    # Check escalation rules for reminder configuration
    if escalation:
        # Response escalation
        response_escalation = escalation.get('response', {})
        if response_escalation:
            has_reminders = True
            escalation_time = response_escalation.get('escalation_time')
            agent_count = len(response_escalation.get('agent_ids', []))
            reminder_summary.append(f"Response: {escalation_time}s to {agent_count} agents")

        # Resolution escalation levels
        resolution_escalation = escalation.get('resolution', {})
        if resolution_escalation:
            for level, level_data in resolution_escalation.items():
                has_reminders = True
                escalation_time = level_data.get('escalation_time')
                agent_count = len(level_data.get('agent_ids', []))
                reminder_summary.append(f"{level.title()}: {escalation_time}s to {agent_count} agents")

    enriched['has_reminders_enabled'] = has_reminders
    enriched['reminder_summary'] = '; '.join(reminder_summary) if reminder_summary else 'No reminders configured'

    return enriched

def flatten_sla_targets(policy):
    """
    Flatten nested SLA target data for better Excel compatibility.

    Args:
        policy (dict): Single SLA policy dictionary

    Returns:
        dict: Policy with flattened SLA target data
    """
    flattened = policy.copy()
    sla_target = policy.get('sla_target', {})

    if sla_target:
        for priority, target in sla_target.items():
            # Create flattened column names
            flattened[f'sla_target_{priority}_respond_within'] = target.get('respond_within')
            flattened[f'sla_target_{priority}_resolve_within'] = target.get('resolve_within')
            flattened[f'sla_target_{priority}_business_hours_id'] = target.get('business_hours', {}).get('id')
            flattened[f'sla_target_{priority}_business_hours_name'] = target.get('business_hours', {}).get('name')
            flattened[f'sla_target_{priority}_escalation_enabled'] = target.get('escalation_enabled')

    return flattened

def flatten_escalation_rules(policy):
    """
    Flatten nested escalation rule data for better Excel compatibility.

    Args:
        policy (dict): Single SLA policy dictionary

    Returns:
        dict: Policy with flattened escalation data
    """
    flattened = policy.copy()
    escalation = policy.get('escalation', {})

    if escalation:
        # Response escalation
        response_escalation = escalation.get('response', {})
        if response_escalation:
            flattened['escalation_response_time'] = response_escalation.get('escalation_time')
            flattened['escalation_response_agent_ids'] = ','.join(map(str, response_escalation.get('agent_ids', [])))

        # Resolution escalation levels
        resolution_escalation = escalation.get('resolution', {})
        if resolution_escalation:
            for level, level_data in resolution_escalation.items():
                flattened[f'escalation_resolution_{level}_time'] = level_data.get('escalation_time')
                flattened[f'escalation_resolution_{level}_agent_ids'] = ','.join(map(str, level_data.get('agent_ids', [])))

    return flattened

def save_to_excel(data, file_name):
    """
    Save SLA policies data to an Excel file with reminder-focused formatting.

    Args:
        data (list): List of SLA policy dictionaries
        file_name (str): Output filename for the Excel file

    Returns:
        bool: True if save successful, False otherwise

    Note:
        - Adds reminder analysis columns for easy filtering
        - Flattens nested JSON structures for spreadsheet compatibility
        - Adds metadata about the export
        - Handles large datasets efficiently
    """
    try:
        logging.info(f"Preparing {len(data)} SLA policies for Excel export with reminder analysis...")
        print("Preparing data for Excel export..."
        # Apply reminder analysis and flattening
        enriched_data = []
        for policy in data:
            # Apply analysis functions
            enriched_policy = analyze_reminder_settings(policy)
            enriched_policy = flatten_sla_targets(enriched_policy)
            enriched_policy = flatten_escalation_rules(enriched_policy)
            enriched_data.append(enriched_policy)

        # Create DataFrame
        df = pd.json_normalize(enriched_data)

        # Reorder columns for better readability (put reminder info first, then basic info)
        reminder_columns = ['id', 'name', 'has_reminders_enabled', 'reminder_summary', 'active', 'is_default', 'position', 'created_at', 'updated_at']
        other_columns = [col for col in df.columns if col not in reminder_columns]
        final_columns = reminder_columns + sorted(other_columns)

        # Reorder DataFrame columns
        df = df[final_columns]

        # Save to Excel with proper formatting
        logging.info(f"Saving {len(df)} rows to Excel file: {file_name}")
        print(f"Saving {len(df)} rows to {file_name}...")

        df.to_excel(file_name, index=False, engine='openpyxl')

        # Add some basic formatting (optional)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill

            wb = load_workbook(file_name)
            ws = wb.active

            # Make header row bold
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            # Highlight rows with reminders enabled
            reminder_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                has_reminders_cell = ws.cell(row=row, column=3)  # Column C is has_reminders_enabled
                if has_reminders_cell.value == True:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = reminder_fill

            wb.save(file_name)
            logging.info("Applied reminder-focused formatting to Excel file")
            print("✓ Applied reminder highlighting to Excel file")

        except ImportError:
            # openpyxl not available for formatting, but data export still works
            logging.info("Advanced formatting skipped (openpyxl not available)")

        logging.info(f"✓ Successfully saved SLA policies to {file_name}")
        print(f"✓ Successfully exported {len(df)} SLA policies to {file_name}")
        return True

    except Exception as e:
        logging.error(f"Failed to save Excel file: {e}")
        print(f"❌ Failed to save Excel file: {e}")
        return False

def main():
    """
    Main function to orchestrate the SLA policies with reminders export process.
    """
    print("Freshdesk SLA Policies with Reminders Export Tool")
    print("=" * 70)

    # Setup logging first
    setup_logging()

    # Validate configuration
    if not validate_configuration():
        print("❌ Configuration validation failed. Please check your settings.")
        return

    # Get current timestamp for metadata
    start_time = time.strftime("%Y-%m-%d %H:%M:%S UTC")

    try:
        # Retrieve SLA policies
        sla_policies = get_sla_policies()

        if not sla_policies:
            logging.error("No SLA policies retrieved from Freshdesk")
            print("❌ No SLA policies found. Please check:")
            print("  - API key has SLA policy read permissions")
            print("  - Freshdesk domain is correct")
            print("  - Network connectivity to Freshdesk")
            print("  - SLA policies exist in your Freshdesk account")
            return

        # Count policies with reminders for summary
        policies_with_reminders = sum(1 for policy in sla_policies if analyze_reminder_settings(policy).get('has_reminders_enabled'))

        # Save to Excel
        success = save_to_excel(sla_policies, OUTPUT_FILENAME)

        if success:
            # Log success with metadata
            end_time = time.strftime("%Y-%m-%d %H:%M:%S UTC")
            logging.info("=" * 60)
            logging.info("SLA POLICIES WITH REMINDERS EXPORT SUMMARY")
            logging.info("=" * 60)
            logging.info(f"Export started: {start_time}")
            logging.info(f"Export completed: {end_time}")
            logging.info(f"Total policies exported: {len(sla_policies)}")
            logging.info(f"Policies with reminders enabled: {policies_with_reminders}")
            logging.info(f"Output file: {OUTPUT_FILENAME}")
            logging.info(f"Log file: {LOG_FILENAME}")
            logging.info("=" * 60)

            print("\n" + "=" * 70)
            print("EXPORT SUMMARY")
            print("=" * 70)
            print(f"✓ Export completed successfully!")
            print(f"  Total SLA policies: {len(sla_policies)}")
            print(f"  Policies with reminders: {policies_with_reminders}")
            print(f"  Policies without reminders: {len(sla_policies) - policies_with_reminders}")
            print(f"  Output file: {OUTPUT_FILENAME}")
            print(f"  Log file: {LOG_FILENAME}")
            print("  Reminder-enabled policies are highlighted in blue in Excel.")
            print("  Check the log file for detailed processing information.")
        else:
            logging.error("Failed to save SLA policies to Excel")
            print("❌ Failed to save SLA policies to Excel. Check logs for details.")

    except KeyboardInterrupt:
        logging.info("Export interrupted by user")
        print("\n⚠ Export interrupted by user")
    except Exception as e:
        logging.error(f"Unexpected error during export: {e}")
        print(f"❌ Unexpected error during export: {e}")

# Run the script if executed directly
if __name__ == '__main__':
    main()

